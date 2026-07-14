#!/usr/bin/env python3
"""Fetch HiFleet weather email and generate dashboard."""
import os
import json
import base64
import re
from datetime import datetime, timedelta
from imap_tools import MailBox, AND, OR
from openpyxl import load_workbook

# === Config ===
EMAIL_USER = os.environ.get("EMAIL_USER", "xuehaifeng666@qq.com")
EMAIL_AUTH = os.environ.get("EMAIL_AUTH", "ztfzkfwzzywmjdaf")
QQ_APP_PWD = os.environ.get("QQ_APP_PWD", "")

VESSEL_LIST_FILE = "Vessel list-20260701.xlsx"
EXCLUDE_SHIPS = {"ORE CHINA", "ORE DONGJIAKOU", "ORE HEBEI", "ORE SHANDONG"}

def get_imap_password():
    return QQ_APP_PWD or EMAIL_AUTH

def fetch_hifleet_email():
    """Fetch latest HiFleet weather email from QQ Mail."""
    password = get_imap_password()
    if not password:
        raise ValueError("No IMAP password available")
    
    with MailBox("imap.qq.com").login(EMAIL_USER, password) as mailbox:
        # Search for hifleet weather emails and get latest one
        criteria = AND(from_="hifleet.com", subject="weather")
        messages = list(mailbox.fetch(criteria, reverse=True, limit=1))
        if messages:
            return messages[0].html
    return None

def load_vessel_list(path):
    """Load vessel list from Excel."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    ships = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell:
                ships.add(str(cell).strip().upper())
    return ships

def parse_hifleet_html(html):
    """Parse HiFleet email HTML to extract ship weather data."""
    ships_data = {}
    
    # Find all ship blocks - match 5 TD cells in a TR
    ship_pattern = re.compile(
        r'<tr[^>]*>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*'
        r'<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>\s*'
        r'<td[^>]*>(.*?)</td>',
        re.DOTALL
    )
    
    for row in ship_pattern.findall(html):
        ship_name = row[0].strip()
        if not ship_name or ship_name.upper() in EXCLUDE_SHIPS:
            continue
        
        forecast = row[1].strip()
        # Parse forecast: format "1d/3.5/1.5/5" or "1d/3.5/1.5"
        # (day_offset, wind, wave, visibility)
        days = re.findall(r'(\d+m?)/(\d+\.?\d*)/(\d+\.?\d*)/(\d+\.?\d*)', forecast)
        
        ships_data[ship_name] = {
            "type": row[2].strip(),
            "forecast": forecast,
            "days": days
        }
    
    return ships_data

def generate_dashboard(ships_data, output_path="output/index.html"):
    """Generate HTML dashboard."""
    os.makedirs("output", exist_ok=True)
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Build alert ships
    alert_ships = []
    for ship, data in ships_data.items():
        if data.get("days") and len(data["days"]) > 0:
            wind = float(data["days"][0][1]) if data["days"][0][1] else 0
            wave = float(data["days"][0][2]) if data["days"][0][2] else 0
            if wind > 6 or wave >= 3:
                alert_ships.append({
                    "ship": ship,
                    "type": data.get("type", ""),
                    "wind": wind,
                    "wave": wave
                })
    
    # Wave >= 3m analysis
    wave3_ships = {}
    for ship, data in ships_data.items():
        wave3_ships[ship] = []
        for i, day in enumerate(data.get("days", [])):
            wave = float(day[2]) if day[2] else 0
            if wave >= 3:
                wave3_ships[ship].append(i + 1)
    
    html = f"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>HiFleet Weather Dashboard - {today}</title>
<style>
body {{ font-family: Arial; background: #0a0e27; color: #c8d3f5; margin: 0; padding: 20px; }}
.header {{ background: linear-gradient(135deg,#1a1f4e,#2d3585); padding:20px; border-radius:12px; margin-bottom:20px; }}
.header h1 {{ color:#7aa2f7; margin:0; }}
.header p {{ color:#7aa2f7; opacity:0.7; margin:5px 0 0; }}
.alert-box {{ background:#1e2340; border-left:4px solid #f7768e; padding:15px; border-radius:8px; margin-bottom:20px; }}
.alert-box h3 {{ color:#f7768e; margin:0 0 10px; }}
table {{ width:100%; border-collapse:collapse; background:#1e2340; border-radius:8px; overflow:hidden; margin-bottom:30px; }}
th {{ background:#7c3aed; color:#fff; padding:10px; text-align:center; }}
td {{ padding:8px 10px; text-align:center; border-bottom:1px solid #2d3585; }}
tr:hover {{ background:#2d3585; }}
.wave-red {{ color:#f7768e; font-weight:bold; }}
.wave-orange {{ color:#ff9e64; }}
.wind-red {{ color:#f7768e; font-weight:bold; }}
section {{ margin-bottom:30px; }}
.ship-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(300px,1fr)); gap:12px; }}
.ship-card {{ background:#1e2340; border-radius:8px; padding:12px; border:1px solid #2d3585; }}
.ship-card h4 {{ color:#7aa2f7; margin:0 0 8px; }}
.badge {{ display:inline-block; padding:2px 8px; border-radius:4px; font-size:12px; margin:2px; }}
.badge-red {{ background:#3d2048; color:#f7768e; }}
.badge-orange {{ background:#3d2a00; color:#ff9e64; }}
</style>
</head>
<body>
<div class="header">
  <h1>&#x1F6A2; HiFleet Weather Dashboard</h1>
  <p>76 ships &middot; Updated: {today} &middot; Alert: wind&gt;6 OR wave&ge;3m</p>
</div>

<div class="alert-box">
  <h3>&#x26A0;&#xFE0F; Today's Alerts ({len(alert_ships)} ships)</h3>
"""
    if alert_ships:
        html += """  <table>
    <tr><th>Ship</th><th>Type</th><th>Wind (m/s)</th><th>Wave (m)</th></tr>
"""
        for s in sorted(alert_ships, key=lambda x: -x["wave"]):
            wind_cls = "wind-red" if s["wind"] > 6 else ""
            wave_cls = "wave-red" if s["wave"] >= 3 else "wave-orange"
            html += f'    <tr><td>{s["ship"]}</td><td>{s["type"]}</td><td class="{wind_cls}">{s["wind"]}</td><td class="{wave_cls}">{s["wave"]}</td></tr>\n'
        html += "  </table>\n"
    else:
        html += "  <p>No alerts today!</p>\n"
    
    html += """</div>

<section>
  <h2 style="color:#7aa2f7;">&#x1F30A; Ships with Wave &ge; 3m (Next 7 Days)</h2>
  <div class="ship-grid">
"""
    
    wave3_count = 0
    for ship, days in sorted(wave3_ships.items()):
        if days:
            wave3_count += 1
            day_labels = {"1": "D1", "2": "D2", "3": "D3", "4": "D4", "5": "D5", "6": "D6", "7": "D7"}
            badges = " ".join([
                f'<span class="badge badge-red">{day_labels.get(str(d), f"D{d}")}</span>'
                for d in days
            ])
            html += f'    <div class="ship-card"><h4>{ship}</h4>{badges}</div>\n'
    
    if wave3_count == 0:
        html += '    <p style="color:#7aa2f7;">No ships with wave &ge; 3m in the next 7 days.</p>\n'
    
    html += """  </div>
</section>

<section>
  <h2 style="color:#7aa2f7;">&#x1F4CA; Full 7-Day Forecast</h2>
  <div style="overflow-x:auto;">
  <table>
    <tr>
      <th>Ship</th><th>Type</th>
      <th>D1 Wind</th><th>D1 Wave</th>
      <th>D2 Wind</th><th>D2 Wave</th>
      <th>D3 Wind</th><th>D3 Wave</th>
      <th>D4 Wind</th><th>D4 Wave</th>
      <th>D5 Wind</th><th>D5 Wave</th>
      <th>D6 Wind</th><th>D6 Wave</th>
      <th>D7 Wind</th><th>D7 Wave</th>
    </tr>
"""
    
    for ship, data in sorted(ships_data.items()):
        days = data.get("days", [])
        row = f'    <tr><td>{ship}</td><td>{data.get("type","")}</td>'
        for i in range(7):
            if i < len(days):
                w = days[i]
                wind = w[1] if w[1] else "-"
                wave = w[2] if w[2] else "-"
                wind_cls = "wind-red" if wind != "-" and float(wind) > 6 else ""
                wave_cls = "wave-red" if wave != "-" and float(wave) >= 3 else ("wave-orange" if wave != "-" and float(wave) >= 2 else "")
                row += f'<td class="{wind_cls}">{wind}</td><td class="{wave_cls}">{wave}</td>'
            else:
                row += "<td>-</td><td>-</td>"
        row += "</tr>\n"
        html += row
    
    html += """  </table>
  </div>
</section>
</body>
</html>"""
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Dashboard saved: {output_path}")

def main():
    print("Fetching HiFleet email...")
    html = fetch_hifleet_email()
    if not html:
        print("No email found, using cached data")
        with open("hifleet_weather_76ships.json") as f:
            ships_data = json.load(f)
    else:
        if os.path.exists(VESSEL_LIST_FILE):
            vessels = load_vessel_list(VESSEL_LIST_FILE)
            print(f"Loaded {len(vessels)} vessels from list")
        else:
            vessels = None
        
        ships_data = parse_hifleet_html(html)
        print(f"Parsed {len(ships_data)} ships from email")
    
    os.makedirs("output", exist_ok=True)
    with open("output/ships_data.json", "w", encoding="utf-8") as f:
        json.dump(ships_data, f, ensure_ascii=False, indent=2)
    
    generate_dashboard(ships_data)
    print("Done!")

if __name__ == "__main__":
    main()
